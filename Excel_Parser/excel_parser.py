"""
Use this program to parse through the NIST management framework
Pandas, openpyxl, and xlrd must be installed to run
Seido Karasaki(yakitategohan on github)
v1 3/23/2023
"""

import openpyxl
import argparse


def get_args():
    with open('description.txt') as f:
        info = f.read()
    parser = argparse.ArgumentParser(
        description='Helps to parse through NIST 800-53A',
        epilog=info
    )
    f.close()
    options = parser.parse_args()
    return options


def read_sheet(link):
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(link)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    return dataframe1


if __name__ == "__main__":
    get_args()
    nist_sheet = read_sheet("NIST-SP800-53r5-Control-Catalog.xlsx")
    while True:
        print('\nHello, would you like to:'
              '\n\n\t1)Search by access control family topic to get a list of control names.'
              '\n\t2)Search by keyword (warning, may return many results if using bad keyword)'
              '\n\t3)Quit')
        option = int(input('\nType 1, 2, or 3: '))
        if option == 1:
            print('\n\nWhat access control family would you like to search for?\n\n'
                  'AC	Access Control	                            PE	Physical and Environmental Protection\n'
                  'AT	Awareness and Training	                    PL	Planning\n'
                  'AU	Audit and Accountability	                PM	Program Management\n'
                  'CA	Assessment, Authorization, and Monitoring	PS	Personnel Security\n'
                  'CM	Configuration Management	                PT	PII Processing and Transparency\n'
                  'CP	Contingency Planning	                    RA	Risk Assessment\n'
                  'IA	Identification and Authentication	        SA	System and Services Acquisition\n'
                  'IR	Incident Response	                        SC	System and Communications Protection\n'
                  'MA	Maintenance	                                SI	System and Information Integrity\n'
                  'MP	Media Protection	                        SR	Supply Chain Risk Management\n')
            topic = input('NIST family to query (please type a two letter title, ex: AC): ')

            # Trap in loop to see if they want to keep searching
            while True:
                # Store hit obj for future use
                hit_obj = {}
                row_count = 1

                # Iterate the loop to read the cell values
                print('\n\n=======================================================================\n\n')
                print('Here are the families that pop up with their control names/identifiers: \n')

                for row_num in range(1, nist_sheet.max_row):
                    cell_obj = nist_sheet.cell(row=row_num, column=1)
                    if cell_obj.value[0:2].lower() == topic.lower():
                        control_name = nist_sheet.cell(row=row_num, column=2)
                        print(f'\t{row_count}) {cell_obj.value} *** {control_name.value}')
                        hit_obj[row_count] = row_num
                        row_count += 1

                print("\n\nWould you like to know further about one of the queries that popped up?\n "
                      "I've saved the row numbers for you. Please type the row number on the left of the control ID"
                      " to get the control text, discussion, or related controls: \n")
                row_to_query = int(input("Type the number here: "))
                while not hit_obj.get(row_to_query):
                    row_to_query = input("Please type valid number")
                while True:
                    print('\n\n=======================================================================\n\n')
                    action = input("What would you like to do with "
                                   f"{nist_sheet.cell(row=hit_obj[row_to_query], column=1).value}:"
                                   "\n\t1)Get control text"
                                   "\n\t2)Get discussion"
                                   "\n\t3)Get related controls"
                                   "\n\t4)Start a new search"
                                   "\n\nPlease type a number: ")
                    action = int(action)
                    # Print control text
                    if action == 1:
                        print(nist_sheet.cell(row=hit_obj[row_to_query], column=3).value)
                    elif action == 2:
                        print(nist_sheet.cell(row=hit_obj[row_to_query], column=4).value)
                    elif action == 3:
                        print(nist_sheet.cell(row=hit_obj[row_to_query], column=5).value)
                    elif action == 4:
                        break
                action = input("Would you like to search from the same list again, or start a new search in a different"
                               " family category? Remember, you can always ctrl+c to quit the program."
                               "\n\t1) Search from same list"
                               "\n\t2) Start a new search"
                               "\n\tPlease enter either 1 or 2: ")
                action = int(action)
                if action == 2:
                    break

        elif option == 2:
            keyword = input("What is the keyword or phrase you would like to search for?\n"
                           "This is not a search engine, so please be specific and use the correct spelling."
                           "\n\n**WARNING** If you search using a generic term, many results may be returned."
                           "\n\nSearch term: ").lower()
            while True:
                hit_obj = {}
                row_count = 1

                # Iterate the loop to read the cell values
                print('\n\n=======================================================================\n\n')
                print('Here is what pops up with that key word/phrase: \n')

                for row_num in range(1, nist_sheet.max_row):
                    id = nist_sheet.cell(row=row_num, column=1).value.lower()
                    name = str(nist_sheet.cell(row=row_num, column=2).value).lower()
                    text = str(nist_sheet.cell(row=row_num, column=3).value).lower()
                    discussion = str(nist_sheet.cell(row=row_num, column=4).value).lower()
                    if keyword in name or keyword in text or keyword in discussion:
                        print(f'\t{row_count}) {id} *** {name}')
                        hit_obj[row_count] = row_num
                        row_count += 1

                print("\n\nWould you like to know further about one of the queries that popped up?\n "
                      "I've saved the row numbers for you. Please type the row number on the left of the control ID"
                      " to get the control text, discussion, or related controls: \n")
                row_to_query = int(input("Type the number here: "))
                while not hit_obj.get(row_to_query):
                    row_to_query = input("Please type valid number")
                while True:
                    print('\n\n=======================================================================\n\n')
                    action = input("What would you like to do with "
                                   f"{nist_sheet.cell(row=hit_obj[row_to_query], column=1).value}:"
                                   "\n\t1)Get control text"
                                   "\n\t2)Get discussion"
                                   "\n\t3)Get related controls"
                                   "\n\t4)Start a new search"
                                   "\n\nPlease type a number: ")
                    action = int(action)
                    # Print control text
                    if action == 1:
                        print(nist_sheet.cell(row=hit_obj[row_to_query], column=3).value)
                    elif action == 2:
                        print(nist_sheet.cell(row=hit_obj[row_to_query], column=4).value)
                    elif action == 3:
                        print(nist_sheet.cell(row=hit_obj[row_to_query], column=5).value)
                    elif action == 4:
                        break
                action = input("Would you like to search from the same list again, or start a new search in a different"
                               " family category? Remember, you can always ctrl+c to quit the program."
                               "\n\t1) Search from same list"
                               "\n\t2) Start a new search"
                               "\n\tPlease enter either 1 or 2: ")
                action = int(action)
                if action == 2:
                    break

        elif option == 3:
            quit(0)

        else:
            print("\n\nPlease type a valid option\n\n")
